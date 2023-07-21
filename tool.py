import openpyxl
import os 
import os.path 
from openpyxl import Workbook

def getdata():
    list_data = []
    list_data1=[]
    wb = openpyxl.load_workbook("F:\Learn-And-Practice\data\data1.XLSX")
    sheet_obj = wb.active
    max_col = sheet_obj.max_column
    # Loop will print all columns name
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj1 = sheet_obj.cell(row = i, column = 2)
        list_data.append(cell_obj.value)
        list_data1.append(cell_obj1.value)
    wb.close()
    return list_data, list_data1
def unitTestcase():
    ut_wb = openpyxl.load_workbook("F:\Learn-And-Practice\data.xlsx")
    ws = ut_wb.create_sheet("Unit TestCase")
    ut_wb.save("F:\Learn-And-Practice\data.xlsx")
    ut_wb.close()

def writeData(data,data2):
    wb=openpyxl.load_workbook("F:\Learn-And-Practice\data.xlsx")
    ws=wb["Unit TestCase"]
    for i in range(len(data)):
        writer_value=ws.cell(row=i+1,column=2)
        writer_value.value=data[i]
        writer_value=ws.cell(row=i+1,column=3)
        writer_value.value=data2[i]
    wb.save("F:\Learn-And-Practice\data.xlsx")
    wb.close()

def main():
    data,data2=getdata()
    unitTestcase()
    writeData(data,data2)
main()